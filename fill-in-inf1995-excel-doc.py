from openpyxl import load_workbook
import os
import logging
import re


class CorrectionException(Exception):
    pass


""" Logger instance for project """
logging.basicConfig(filename='output.log', filemode='w')
console = logging.StreamHandler()
logger = logging.getLogger('Excel helper')
logger.addHandler(console)

FIRST_ROW = 11
LAST_ROW = 16
GROUP_COLUMN_LETTER = 'P'
GRADE_COLUMN_LETTER = 'K'
CORRECTION_FILE_NAME = 'Correction_tp8.txt'
CORRECTION_PATH = os.path.join('test', 'section_02')
GRADES_WORKBOOK = 'notes-inf1995-sect02-test.xlsx'
OUTPUT_WORKBOOK = 'notes-inf1995-sect02-test-rempli.xlsx'

GROUP_COLUMN = ord(GROUP_COLUMN_LETTER) - ord('A') + 1
GRADE_COLUMN = ord(GRADE_COLUMN_LETTER) - ord('A') + 1


def fill_in_grades(workbook, output_name):
    """
    Fill in all the rows from FIRST_ROW to LAST_ROW inclusively
    """
    wb = load_workbook(workbook)
    s = wb.get_active_sheet()
    for row_number in range(FIRST_ROW, LAST_ROW+1):
        try:
            fill_in_row(s, row_number)
        except CorrectionException as e:
            logger.warning(
                "could not fill in row_number={}".format(row_number)
                + '\n   ' + str(e).replace('\n', '\n   ')
            )

    wb.save(output_name)


def fill_in_row(sheet, row_number):
    """
    Fill in a single row of the excel sheet
    """
    total = 'SEE_LOG'
    group = get_group(sheet, row_number)
    try:
        int(group)
        total = get_total(group)
    except ValueError:
        raise CorrectionException(
            "Invalid group={}"
            .format(group.encode('ascii', 'replace'))
        )
    finally:
        set_total(sheet, row_number, total)


def get_total(group):
    """
    Gets the total points in the group's corresponding correction file
    """
    path = make_path_from_group(group)
    value = get_value_from_file(path)
    return value


def make_path_from_group(group):
    """ 
    Constructs a path to the correction file from the group number
    """
    path = os.path.join(CORRECTION_PATH, str(group), CORRECTION_FILE_NAME)
    return path


def get_value_from_file(path):
    """ 
    Extracts the total number of points from the file specified by path
    """
    try:
        with open(path, 'r', encoding='UTF-8') as f:
            return get_value_from_filecontent(f.read())
    except FileNotFoundError:
        raise CorrectionException('File not found : {}'.format(path))
    except CorrectionException as e:
        raise CorrectionException("Could not find total in file={}".format(path) + '\n' + str(e))


def get_value_from_filecontent(file_content):
    """ 
    Extracts the number of points from the number of points from the
    content of the file by isolating the line containing the total and
    extracting the total from that line
    """
    total_line_re = re.compile(r'Total:.*', re.MULTILINE)
    try:
        matches = total_line_re.findall(file_content)
        total_line = matches[0]
    except IndexError:
        raise CorrectionException("No line matching total line regex={}".format(total_line_re))

    try:
        total_re = re.compile(r'[1-9][0-9]?')
        total = total_re.findall(total_line)[0]
    except IndexError:
        raise CorrectionException("Could not find points on total_line={}".format(total_line))

    return total


def get_group(sheet, row_number):
    """
    Gets the group from a row_number in the worksheet
    """
    group = sheet.cell(row=row_number, column=GROUP_COLUMN).value
    if group is None:
        raise CorrectionException("No group in cell row={}, column={}"
                                  .format(row_number, GROUP_COLUMN_LETTER))
    return group


def set_total(sheet, row_number, total):
    """
    Set the total in the worksheet
    """
    sheet.cell(row=row_number, column=GRADE_COLUMN).value = total


if __name__ == "__main__":
    fill_in_grades(GRADES_WORKBOOK, OUTPUT_WORKBOOK)
